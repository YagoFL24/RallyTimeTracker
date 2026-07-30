import csv
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from gestorTiempos import MAX_SQLITE_INTEGER, milisegundos_a_tiempo, tiempo_a_milisegundos


FORMAT_NAME = "RallyTimeTracker"
FORMAT_VERSION = 1
DATA_HEADERS = [
    "formato",
    "version",
    "competicion",
    "fecha",
    "numero_tramos",
    "participante",
    "estado_rally",
    "retirado_tras_tramo",
    "estado_previo_descalificacion",
    "tramo",
    "estado_tramo",
    "tiempo",
    "tiempo_anterior",
    "revisiones",
    "actualizado_en",
]
RALLY_STATUS_LABELS = {
    "active": "Activo",
    "retired": "Retirado",
    "disqualified": "Descalificado",
}
RALLY_STATUS_VALUES = {
    **{label.casefold(): value for value, label in RALLY_STATUS_LABELS.items()},
    **{value: value for value in RALLY_STATUS_LABELS},
}
STAGE_STATUS_LABELS = {
    "pending": "Pendiente",
    "finished": "Finalizado",
    "stage_dnf": "No finalizado",
    "dns": "No presentado",
    "dsq": "Descalificado",
}
STAGE_STATUS_VALUES = {
    **{label.casefold(): value for value, label in STAGE_STATUS_LABELS.items()},
    **{value: value for value in STAGE_STATUS_LABELS},
}


class ExchangeError(ValueError):
    pass


def _display_time(milliseconds):
    if milliseconds is None:
        return ""
    return milisegundos_a_tiempo(milliseconds)


def _data_rows(competition):
    records = {
        row["participant_name"]: row
        for row in competition["participant_records"]
    }
    results = {
        (row["participant_name"], row["stage_number"]): row
        for row in competition["results"]
    }
    rows = []
    for participant in competition["participants"]:
        record = records[participant]
        for stage in range(1, competition["stages"] + 1):
            result = results[(participant, stage)]
            rows.append(
                {
                    "formato": FORMAT_NAME,
                    "version": FORMAT_VERSION,
                    "competicion": competition["name"],
                    "fecha": competition.get("event_date") or "",
                    "numero_tramos": competition["stages"],
                    "participante": participant,
                    "estado_rally": RALLY_STATUS_LABELS[record["rally_status"]],
                    "retirado_tras_tramo": record["retired_after_stage"] or "",
                    "estado_previo_descalificacion": RALLY_STATUS_LABELS.get(
                        record.get("status_before_disqualification"), ""
                    ),
                    "tramo": stage,
                    "estado_tramo": STAGE_STATUS_LABELS[result["status"]],
                    "tiempo": _display_time(result["time_ms"]),
                    "tiempo_anterior": _display_time(result["previous_time_ms"]),
                    "revisiones": result["revision_count"],
                    "actualizado_en": result["updated_at"],
                }
            )
    return rows


def export_data(competition, destination):
    path = Path(destination)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        _write_csv(path, competition)
    elif suffix == ".xlsx":
        _write_excel(path, competition)
    else:
        raise ExchangeError("El formato debe ser CSV (.csv) o Excel (.xlsx).")


def _write_csv(path, competition):
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=DATA_HEADERS, delimiter=";")
        writer.writeheader()
        writer.writerows(
            {
                header: _csv_safe(row[header])
                for header in DATA_HEADERS
            }
            for row in _data_rows(competition)
        )


def _csv_safe(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _write_excel(path, competition):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ExchangeError("Falta la dependencia openpyxl para crear Excel.") from exc

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Datos"
    data_sheet.append(DATA_HEADERS)
    for row in _data_rows(competition):
        data_sheet.append([row[header] for header in DATA_HEADERS])
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions

    classification = workbook.create_sheet("Clasificación")
    classification_headers = (
        ["Pos", "Piloto", "Estado"]
        + [f"Tramo {stage}" for stage in range(1, competition["stages"] + 1)]
        + ["Tramos ganados", "General", "Dif."]
    )
    classification.append(classification_headers)
    for row in competition["leaderboard"]:
        difference = ""
        if row["diff"] not in (None, 0):
            difference = f"+{_display_time(row['diff'])}"
        classification.append(
            [
                row["rank"],
                row["participant"],
                row["classification_status"],
                *[_format_stage_result(result) for result in row["stage_results"]],
                row["stage_wins"],
                _display_time(row["total"]),
                difference,
            ]
        )
    classification.freeze_panes = "A2"
    classification.auto_filter.ref = classification.dimensions

    for sheet in (data_sheet, classification):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                    cell.data_type = "s"

    header_fill = PatternFill("solid", fgColor="0E639C")
    for sheet in (data_sheet, classification):
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
            width = min(max(len(value) for value in values) + 2, 32)
            sheet.column_dimensions[get_column_letter(column)].width = max(width, 10)

    workbook.save(path)


def _format_stage_result(result):
    status = result["status"]
    if status == "finished":
        return _display_time(result["time_ms"])
    if status == "stage_dnf":
        return f"NF {_display_time(result['time_ms'])}"
    if status == "dns":
        return "NP"
    if status == "dsq":
        return "DSQ"
    return "Pendiente" if result.get("stage_started") else "-"


def read_data(source):
    path = Path(source)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        rows = _read_csv(path)
    elif suffix == ".xlsx":
        rows = _read_excel(path)
    else:
        raise ExchangeError("Solo se pueden importar archivos CSV o Excel.")
    return _rows_to_snapshot(rows)


def _read_csv(path):
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            sample = file.read(4096)
            file.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ";"
            return list(csv.DictReader(file, delimiter=delimiter))
    except UnicodeDecodeError as exc:
        raise ExchangeError("El CSV debe estar codificado como UTF-8.") from exc


def _read_excel(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExchangeError("Falta la dependencia openpyxl para leer Excel.") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExchangeError("El archivo Excel está dañado o no es compatible.") from exc
    try:
        sheet = workbook["Datos"] if "Datos" in workbook.sheetnames else workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if headers is None:
            return []
        normalized_headers = [str(value or "").strip() for value in headers]
        if len(set(normalized_headers)) != len(normalized_headers):
            raise ExchangeError("El archivo Excel contiene columnas duplicadas.")
        return [dict(zip(normalized_headers, values)) for values in iterator]
    finally:
        workbook.close()


def _text(value):
    text = "" if value is None else str(value).strip()
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _integer(value, field, minimum=0, maximum=MAX_SQLITE_INTEGER):
    if isinstance(value, bool):
        raise ExchangeError(f"El campo '{field}' debe ser un entero.")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ExchangeError(f"El campo '{field}' debe ser un entero.") from exc
    if str(number) != _text(value) and not isinstance(value, (int, float)):
        raise ExchangeError(f"El campo '{field}' debe ser un entero.")
    if number < minimum:
        raise ExchangeError(f"El campo '{field}' debe ser al menos {minimum}.")
    if number > maximum:
        raise ExchangeError(f"El campo '{field}' es demasiado grande.")
    return number


def _optional_integer(value, field, minimum=0):
    return None if not _text(value) else _integer(value, field, minimum)


def _status(value, mapping, field):
    key = _text(value).casefold()
    if key not in mapping:
        raise ExchangeError(f"Valor desconocido en '{field}': {_text(value)}.")
    return mapping[key]


def _time(value, field):
    text = _text(value)
    if not text:
        return None
    try:
        return tiempo_a_milisegundos(text)
    except (TypeError, ValueError) as exc:
        raise ExchangeError(f"Tiempo inválido en '{field}': {text}.") from exc


def _rows_to_snapshot(rows):
    if not rows:
        raise ExchangeError("El archivo no contiene resultados.")
    missing = [header for header in DATA_HEADERS if header not in rows[0]]
    if missing:
        raise ExchangeError(f"Faltan columnas obligatorias: {', '.join(missing)}.")

    first = rows[0]
    if _text(first["formato"]) != FORMAT_NAME:
        raise ExchangeError("El archivo no pertenece a RallyTimeTracker.")
    version = _integer(first["version"], "version", 1)
    if version != FORMAT_VERSION:
        raise ExchangeError(f"La versión de intercambio {version} no es compatible.")
    competition_name = _text(first["competicion"])
    if not competition_name:
        raise ExchangeError("El nombre de la competición está vacío.")
    stages = _integer(first["numero_tramos"], "numero_tramos", 1)
    event_date = _text(first["fecha"]) or None

    participants = {}
    for row_number, row in enumerate(rows, start=2):
        if _text(row.get("formato")) != FORMAT_NAME:
            raise ExchangeError(f"Formato incoherente en la fila {row_number}.")
        if _integer(row.get("version"), "version", 1) != version:
            raise ExchangeError(f"Versión incoherente en la fila {row_number}.")
        if _text(row.get("competicion")) != competition_name:
            raise ExchangeError(f"Competición incoherente en la fila {row_number}.")
        if _integer(row.get("numero_tramos"), "numero_tramos", 1) != stages:
            raise ExchangeError(f"Número de tramos incoherente en la fila {row_number}.")
        if (_text(row.get("fecha")) or None) != event_date:
            raise ExchangeError(f"Fecha incoherente en la fila {row_number}.")

        name = _text(row.get("participante"))
        if not name:
            raise ExchangeError(f"Participante vacío en la fila {row_number}.")
        rally_status = _status(row.get("estado_rally"), RALLY_STATUS_VALUES, "estado_rally")
        retired_after = _optional_integer(
            row.get("retirado_tras_tramo"), "retirado_tras_tramo", 1
        )
        before_text = _text(row.get("estado_previo_descalificacion"))
        status_before = (
            _status(before_text, RALLY_STATUS_VALUES, "estado_previo_descalificacion")
            if before_text
            else None
        )
        stage = _integer(row.get("tramo"), "tramo", 1)
        if stage > stages:
            raise ExchangeError(f"Tramo fuera de rango en la fila {row_number}.")
        stage_status = _status(row.get("estado_tramo"), STAGE_STATUS_VALUES, "estado_tramo")
        time_ms = _time(row.get("tiempo"), "tiempo")
        previous_time_ms = _time(row.get("tiempo_anterior"), "tiempo_anterior")
        revision_count = _integer(row.get("revisiones"), "revisiones", 0)
        if stage_status in {"finished", "stage_dnf"} and time_ms is None:
            raise ExchangeError(f"Falta el tiempo en la fila {row_number}.")
        if stage_status in {"pending", "dns", "dsq"} and time_ms is not None:
            raise ExchangeError(f"El estado no admite tiempo en la fila {row_number}.")

        participant = participants.setdefault(
            name.casefold(),
            {
                "name": name,
                "rally_status": rally_status,
                "retired_after_stage": retired_after,
                "status_before_disqualification": status_before,
                "results": {},
            },
        )
        metadata = (rally_status, retired_after, status_before)
        expected = (
            participant["rally_status"],
            participant["retired_after_stage"],
            participant["status_before_disqualification"],
        )
        if metadata != expected:
            raise ExchangeError(f"Estado incoherente para {name} en la fila {row_number}.")
        if stage in participant["results"]:
            raise ExchangeError(f"Tramo duplicado para {name} en la fila {row_number}.")
        participant["results"][stage] = {
            "stage_number": stage,
            "status": stage_status,
            "time_ms": time_ms,
            "previous_time_ms": previous_time_ms,
            "revision_count": revision_count,
            "updated_at": _text(row.get("actualizado_en"))
            or datetime.now().isoformat(sep=" ", timespec="seconds"),
        }

    normalized = []
    for participant in participants.values():
        if set(participant["results"]) != set(range(1, stages + 1)):
            raise ExchangeError(
                f"{participant['name']} no contiene exactamente los {stages} tramos."
            )
        retired_state = participant["rally_status"] == "retired" or (
            participant["rally_status"] == "disqualified"
            and participant["status_before_disqualification"] == "retired"
        )
        if retired_state:
            retired_after = participant["retired_after_stage"]
            if retired_after is None or retired_after > stages:
                raise ExchangeError(
                    f"La retirada de {participant['name']} no indica un tramo válido."
                )
        elif participant["retired_after_stage"] is not None:
            raise ExchangeError(
                f"Solo un participante retirado puede indicar el tramo de retirada."
            )
        has_dsq = any(
            row["status"] == "dsq" for row in participant["results"].values()
        )
        if participant["rally_status"] == "disqualified":
            if participant["status_before_disqualification"] not in {"active", "retired"}:
                raise ExchangeError(
                    f"Falta el estado anterior a la descalificación de {participant['name']}."
                )
        elif has_dsq:
            raise ExchangeError(
                f"El participante {participant['name']} tiene un tramo DSQ pero no está descalificado."
            )
        participant["results"] = [
            participant["results"][stage] for stage in range(1, stages + 1)
        ]
        normalized.append(participant)

    return {
        "name": competition_name,
        "stages": stages,
        "event_date": event_date,
        "participants": normalized,
    }


def export_pdf(competition, destination):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise ExchangeError("Falta la dependencia reportlab para crear PDF.") from exc

    path = Path(destination)
    if path.suffix.casefold() != ".pdf":
        raise ExchangeError("La clasificación imprimible debe guardarse como PDF.")
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Clasificación - {competition['name']}",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Clasificación — {escape(competition['name'])}", styles["Title"]),
        Paragraph(f"Tramos: {competition['stages']}", styles["Normal"]),
        Spacer(1, 5 * mm),
    ]
    chunks = [
        list(range(start, min(start + 8, competition["stages"] + 1)))
        for start in range(1, competition["stages"] + 1, 8)
    ]
    for chunk_index, stage_numbers in enumerate(chunks):
        if chunk_index:
            story.append(PageBreak())
            story.append(Paragraph(f"Clasificación — {escape(competition['name'])}", styles["Heading1"]))
        headers = ["Pos", "Piloto", "Estado"] + [f"T{stage}" for stage in stage_numbers] + ["Tramos ganados", "General", "Dif."]
        table_rows = [headers]
        for row in competition["leaderboard"]:
            difference = "-" if row["diff"] in (None, 0) else f"+{_display_time(row['diff'])}"
            table_rows.append(
                [row["rank"], row["participant"], row["classification_status"]]
                + [_format_stage_result(row["stage_results"][stage - 1]) for stage in stage_numbers]
                + [row["stage_wins"], _display_time(row["total"]), difference]
            )
        available_width = landscape(A4)[0] - 20 * mm
        fixed_width = 10 * mm + 45 * mm + 30 * mm + 25 * mm + 25 * mm + 25 * mm
        stage_width = (available_width - fixed_width) / max(len(stage_numbers), 1)
        table = Table(
            table_rows,
            repeatRows=1,
            colWidths=[10 * mm, 45 * mm, 30 * mm]
            + [stage_width] * len(stage_numbers)
            + [25 * mm, 25 * mm, 25 * mm],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E639C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (3, 1), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2F8")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(table)
    document.build(story)


def _championship_headers(championship, event_indexes):
    return [
        "Pos",
        "Piloto",
        "Estado",
        "Puntos",
        "Diferencia",
        "Victorias",
        "Podios",
        "Tramos ganados",
        "Abandonos",
    ] + [championship["events"][index]["competition_name"] for index in event_indexes]


def _championship_row(row, event_indexes):
    event_points = []
    for index in event_indexes:
        result = row["event_results"][index]
        event_points.append("-" if result is None else result["total_points"])
    return [
        row["rank"],
        row["driver"],
        "Activo" if row["status"] == "active" else "Retirado",
        row["points"],
        row["difference"],
        row["wins"],
        row["podiums"],
        row["stage_wins"],
        row["retirements"],
        *event_points,
    ]


def export_championship_data(championship, destination):
    path = Path(destination)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        _write_championship_csv(path, championship)
    elif suffix == ".xlsx":
        _write_championship_excel(path, championship)
    else:
        raise ExchangeError("El campeonato debe exportarse como .csv o .xlsx.")


def _write_championship_csv(path, championship):
    event_indexes = list(range(len(championship["events"])))
    try:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, delimiter=";", lineterminator="\n")
            writer.writerow(_championship_headers(championship, event_indexes))
            for row in championship["standings"]:
                writer.writerow(
                    [_csv_safe(value) for value in _championship_row(row, event_indexes)]
                )
    except OSError as exc:
        raise ExchangeError(f"No se pudo escribir el CSV: {exc}") from exc


def _write_championship_excel(path, championship):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ExchangeError("Falta la dependencia openpyxl para crear Excel.") from exc
    workbook = Workbook()
    standings_sheet = workbook.active
    standings_sheet.title = "Clasificacion"
    event_indexes = list(range(len(championship["events"])))
    standings_sheet.append(_championship_headers(championship, event_indexes))
    for row in championship["standings"]:
        standings_sheet.append(_championship_row(row, event_indexes))

    calendar_sheet = workbook.create_sheet("Calendario")
    calendar_sheet.append(["Orden", "Competicion", "Fecha", "Estado"])
    for event in championship["events"]:
        calendar_sheet.append(
            [
                event["event_order"],
                _csv_safe(event["competition_name"]),
                event["event_date"] or "",
                event["status"],
            ]
        )

    points_sheet = workbook.create_sheet("Puntuacion")
    points_sheet.append(["Posicion", "Puntos"])
    for position, points in enumerate(championship["points_table"], start=1):
        points_sheet.append([position, points])
    points_sheet.append([])
    points_sheet.append(["Bonificacion por victorias de tramo", championship["stage_win_bonus"]])

    header_fill = PatternFill("solid", fgColor="0E639C")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
            sheet.column_dimensions[column[0].column_letter].width = width
    try:
        workbook.save(path)
    except OSError as exc:
        raise ExchangeError(f"No se pudo escribir el Excel: {exc}") from exc


def export_championship_pdf(championship, destination):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise ExchangeError("Falta la dependencia reportlab para crear PDF.") from exc
    path = Path(destination)
    if path.suffix.casefold() != ".pdf":
        raise ExchangeError("La clasificacion del campeonato debe guardarse como PDF.")
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Campeonato - {championship['name']}",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Campeonato — {escape(championship['name'])}", styles["Title"]),
        Paragraph(
            f"Estado: {championship['status']} · Bonificación por tramos: "
            f"{championship['stage_win_bonus']} puntos",
            styles["Normal"],
        ),
        Spacer(1, 4 * mm),
    ]
    chunks = [
        list(range(start, min(start + 5, len(championship["events"]))))
        for start in range(0, len(championship["events"]), 5)
    ] or [[]]
    for chunk_number, event_indexes in enumerate(chunks):
        if chunk_number:
            story.append(PageBreak())
            story.append(
                Paragraph(
                    f"Campeonato — {escape(championship['name'])}",
                    styles["Heading1"],
                )
            )
        rows = [_championship_headers(championship, event_indexes)]
        rows.extend(
            _championship_row(row, event_indexes)
            for row in championship["standings"]
        )
        table = Table(rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E639C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2F8")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(table)
    document.build(story)
