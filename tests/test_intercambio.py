import csv
import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from servicios import RallyService


class ExchangeTests(unittest.TestCase):
    def setUp(self):
        self.original_cwd = os.getcwd()
        self.temp_dir = tempfile.TemporaryDirectory()
        os.chdir(self.temp_dir.name)
        self.service = RallyService()
        ok, message = self.service.create_competition(
            "Rally Ártico", 2, ["Ana", "Luis", "Marta"]
        )
        self.assertTrue(ok, message)
        for participant, stage, time_text in (
            ("Ana", 1, "1:00.000"),
            ("Ana", 1, "1:01.250"),
            ("Luis", 1, "1:02.000"),
            ("Ana", 2, "1:30.000"),
        ):
            self.assertTrue(
                self.service.add_time_str(
                    "Rally Ártico", participant, stage, time_text
                )[0]
            )
        self.assertTrue(
            self.service.set_result_status(
                "Rally Ártico", "Luis", 2, "No finalizado"
            )[0]
        )
        self.assertTrue(
            self.service.retire_from_rally(
                "Rally Ártico", "Luis", 2, True
            )[0]
        )
        self.assertTrue(
            self.service.set_result_status(
                "Rally Ártico", "Marta", 1, "Descalificado"
            )[0]
        )

    def tearDown(self):
        os.chdir(self.original_cwd)
        self.temp_dir.cleanup()

    def test_csv_round_trip_preserves_results_and_uses_new_name(self):
        destination = Path("rally.csv")
        self.assertTrue(
            self.service.export_competition("Rally Ártico", destination)[0]
        )
        raw = destination.read_text(encoding="utf-8-sig")
        self.assertIn("competicion;fecha;numero_tramos", raw)
        self.assertIn("Rally Ártico", raw)

        ok, message, imported_name = self.service.import_competition(destination)
        self.assertTrue(ok, message)
        self.assertEqual(imported_name, "Rally Ártico_importada")
        imported = self.service.get_competition_info(imported_name)
        records = {
            row["participant_name"]: row
            for row in imported["participant_records"]
        }
        self.assertEqual(records["Luis"]["rally_status"], "retired")
        self.assertEqual(records["Luis"]["retired_after_stage"], 2)
        self.assertEqual(records["Marta"]["rally_status"], "disqualified")
        results = {
            (row["participant_name"], row["stage_number"]): row
            for row in imported["results"]
        }
        self.assertEqual(results[("Ana", 1)]["time_ms"], 61_250)
        self.assertEqual(results[("Ana", 1)]["previous_time_ms"], 60_000)
        self.assertEqual(results[("Ana", 1)]["revision_count"], 1)
        self.assertEqual(results[("Luis", 2)]["status"], "stage_dnf")
        self.assertEqual(results[("Marta", 1)]["status"], "dsq")

    def test_excel_contains_data_and_classification_and_can_be_imported(self):
        destination = Path("rally.xlsx")
        self.assertTrue(
            self.service.export_competition("Rally Ártico", destination)[0]
        )
        workbook = load_workbook(destination, read_only=True, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["Datos", "Clasificación"])
            self.assertEqual(workbook["Datos"]["A2"].value, "RallyTimeTracker")
            self.assertEqual(workbook["Clasificación"]["A1"].value, "Pos")
            headers = [
                cell.value for cell in next(workbook["Clasificación"].iter_rows())
            ]
            self.assertIn("Tramos ganados", headers)
        finally:
            workbook.close()

        self.assertTrue(self.service.import_competition(destination)[0])
        ok, message, imported_name = self.service.import_competition(destination)
        self.assertTrue(ok, message)
        self.assertEqual(imported_name, "Rally Ártico_importada_2")

    def test_invalid_file_is_rejected_without_creating_competition(self):
        destination = Path("invalido.csv")
        with destination.open("w", encoding="utf-8", newline="") as file:
            writer = csv.writer(file, delimiter=";")
            writer.writerow(["formato", "version"])
            writer.writerow(["Otro", "1"])
        before = self.service.list_competitions()
        ok, message, imported_name = self.service.import_competition(destination)
        self.assertFalse(ok)
        self.assertIn("Faltan columnas", message)
        self.assertIsNone(imported_name)
        self.assertEqual(self.service.list_competitions(), before)

    def test_damaged_excel_is_reported_without_crashing(self):
        destination = Path("dañado.xlsx")
        destination.write_bytes(b"esto no es un libro excel")
        ok, message, imported_name = self.service.import_competition(destination)
        self.assertFalse(ok)
        self.assertIn("dañado", message)
        self.assertIsNone(imported_name)

    def test_excel_keeps_formula_like_participant_as_plain_text(self):
        ok, message = self.service.create_competition("Fórmulas", 1, ["=PILOTO"])
        self.assertTrue(ok, message)
        destination = Path("formulas.xlsx")
        self.assertTrue(self.service.export_competition("Fórmulas", destination)[0])
        workbook = load_workbook(destination, read_only=False, data_only=False)
        try:
            participant_cell = workbook["Datos"]["F2"]
            self.assertEqual(participant_cell.value, "=PILOTO")
            self.assertEqual(participant_cell.data_type, "s")
        finally:
            workbook.close()
        self.assertTrue(self.service.import_competition(destination)[0])
        imported = self.service.get_competition_info("Fórmulas_importada")
        self.assertEqual(imported["participants"], ["=PILOTO"])

    def test_pdf_is_generated_with_a_valid_signature(self):
        destination = Path("clasificacion.pdf")
        ok, message = self.service.export_classification_pdf(
            "Rally Ártico", destination
        )
        self.assertTrue(ok, message)
        self.assertTrue(destination.read_bytes().startswith(b"%PDF-"))
        self.assertGreater(destination.stat().st_size, 1_000)


if __name__ == "__main__":
    unittest.main()
