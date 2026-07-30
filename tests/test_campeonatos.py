import os
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from database_schema import CORE_SCHEMA_STATEMENTS, SCHEMA_VERSION
from persistencia import get_competition, start_connection
from servicios import RallyService


class ChampionshipTestCase(unittest.TestCase):
    def setUp(self):
        self.original_cwd = os.getcwd()
        self.temp_dir = tempfile.TemporaryDirectory()
        os.chdir(self.temp_dir.name)
        self.service = RallyService()

    def tearDown(self):
        os.chdir(self.original_cwd)
        self.temp_dir.cleanup()

    def create_championship(self, drivers=None, bonus=5):
        drivers = drivers or ["Ana", "Luis", "Marta"]
        ok, message = self.service.create_championship(
            "Regional", drivers, stage_win_bonus=bonus
        )
        self.assertTrue(ok, message)

    def create_competition(self, name="Rally Norte", participants=None, stages=2):
        participants = participants or ["Ana", "Luis", "Marta"]
        ok, message = self.service.create_competition(name, stages, participants)
        self.assertTrue(ok, message)

    def finish_competition(self, name, participants, stage_times):
        for participant in participants:
            for stage, time_text in enumerate(stage_times[participant], start=1):
                ok, message = self.service.add_time_str(
                    name, participant, stage, time_text
                )
                self.assertTrue(ok, message)


class ChampionshipSchemaTests(ChampionshipTestCase):
    def test_migrates_v2_database_without_losing_competitions(self):
        data_dir = Path("data")
        data_dir.mkdir()
        database_path = data_dir / "datos.db"
        connection = sqlite3.connect(database_path)
        for statement in CORE_SCHEMA_STATEMENTS:
            connection.execute(statement)
        connection.execute(
            "INSERT INTO competitions (competition_name, number_of_stages) "
            "VALUES ('Existente', 1)"
        )
        connection.execute("PRAGMA user_version = 2")
        connection.commit()
        connection.close()

        migrated, _cursor = start_connection()
        tables = {
            row[0]
            for row in migrated.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        self.assertEqual(
            migrated.execute("PRAGMA user_version").fetchone()[0], SCHEMA_VERSION
        )
        migrated.close()

        self.assertIsNotNone(get_competition("Existente"))
        self.assertTrue(
            {
                "drivers",
                "driver_aliases",
                "championships",
                "championship_events",
            }.issubset(tables)
        )
        self.assertTrue((data_dir / "datos.v2.backup.db").is_file())


class ChampionshipLifecycleTests(ChampionshipTestCase):
    def test_creates_championship_and_attaches_existing_competition_with_aliases(self):
        self.create_championship(["Ana Oficial", "Luis Oficial"])
        self.create_competition(participants=["Ana", "Luis", "Invitado"])

        ok, message = self.service.add_competition_to_championship(
            "Regional",
            "Rally Norte",
            {"Ana Oficial": "Ana", "Luis Oficial": "Luis"},
        )
        self.assertTrue(ok, message)

        championship = self.service.get_championship_info("Regional")
        self.assertEqual(len(championship["events"]), 1)
        self.assertEqual(
            {row["official_name"]: row["aliases"] for row in championship["drivers"]},
            {
                "Ana Oficial": ["Ana", "Ana Oficial"],
                "Luis Oficial": ["Luis", "Luis Oficial"],
            },
        )
        mappings = championship["events"][0]["driver_mappings"]
        self.assertEqual(
            {row["participant_name"] for row in mappings}, {"Ana", "Luis"}
        )

    def test_competition_can_belong_to_multiple_championships(self):
        self.create_competition(participants=["Ana", "Luis"])
        for championship_name in ("Regional", "Nacional"):
            self.assertTrue(
                self.service.create_championship(
                    championship_name, ["Ana", "Luis"]
                )[0]
            )
            self.assertTrue(
                self.service.add_competition_to_championship(
                    championship_name, "Rally Norte"
                )[0]
            )

        ok, message = self.service.delete_competition("Rally Norte")
        self.assertFalse(ok)
        self.assertIn("Nacional", message)
        self.assertIn("Regional", message)

    def test_created_championship_competition_uses_active_roster(self):
        self.create_championship()
        ok, message = self.service.create_competition_for_championship(
            "Regional", "Rally Nuevo", 3, "2026-09-12"
        )
        self.assertTrue(ok, message)

        competition = self.service.get_competition_info("Rally Nuevo")
        self.assertEqual(competition["participants"], ["Ana", "Luis", "Marta"])
        championship = self.service.get_championship_info("Regional")
        self.assertEqual(championship["events"][0]["competition_name"], "Rally Nuevo")
        self.assertEqual(championship["events"][0]["event_date"], "2026-09-12")

    def test_removes_and_reorders_events_without_deleting_competitions(self):
        self.create_championship()
        for name in ("Primera", "Segunda", "Tercera"):
            self.create_competition(name)
            self.assertTrue(
                self.service.add_competition_to_championship("Regional", name)[0]
            )
        championship = self.service.get_championship_info("Regional")
        second_id = championship["events"][1]["id"]
        self.assertTrue(
            self.service.move_competition_in_championship(
                "Regional", second_id, -1
            )[0]
        )
        championship = self.service.get_championship_info("Regional")
        self.assertEqual(
            [event["competition_name"] for event in championship["events"]],
            ["Segunda", "Primera", "Tercera"],
        )
        first_id = championship["events"][1]["id"]
        self.assertTrue(
            self.service.remove_competition_from_championship(
                "Regional", first_id
            )[0]
        )
        self.assertIsNotNone(self.service.get_competition_info("Primera"))
        championship = self.service.get_championship_info("Regional")
        self.assertEqual(
            [event["event_order"] for event in championship["events"]], [1, 2]
        )

    def test_rejects_event_when_an_active_driver_cannot_be_mapped(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Invitado"])

        ok, message = self.service.add_competition_to_championship(
            "Regional", "Rally Norte"
        )
        self.assertFalse(ok)
        self.assertIn("Luis", message)

    def test_deleting_championship_preserves_competition_and_creates_backup(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"])
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )

        ok, message = self.service.delete_championship("Regional")
        self.assertTrue(ok, message)
        self.assertIsNotNone(self.service.get_competition_info("Rally Norte"))
        self.assertIsNone(self.service.get_championship_info("Regional"))
        backups, _directory, error = self.service.list_database_backups()
        self.assertIsNone(error)
        self.assertEqual(backups[0]["reason"], "pre_championship")


class ChampionshipScoringTests(ChampionshipTestCase):
    def test_scores_registered_drivers_and_stage_win_bonus_ignoring_guests(self):
        self.create_championship()
        participants = ["Ana", "Luis", "Marta", "Invitado"]
        self.create_competition(participants=participants)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.finish_competition(
            "Rally Norte",
            participants,
            {
                "Ana": ["1:00.000", "1:05.000"],
                "Luis": ["1:00.000", "1:04.000"],
                "Marta": ["1:10.000", "1:06.000"],
                "Invitado": ["0:50.000", "0:50.000"],
            },
        )

        championship = self.service.get_championship_info("Regional")
        self.assertEqual(championship["status"], "Finalizado")
        standings = {row["driver"]: row for row in championship["standings"]}
        self.assertEqual(standings["Luis"]["points"], 30)
        self.assertEqual(standings["Luis"]["stage_wins"], 2)
        self.assertEqual(standings["Ana"]["points"], 18)
        self.assertEqual(standings["Ana"]["stage_wins"], 1)
        self.assertEqual(standings["Marta"]["points"], 15)

    def test_tied_rally_and_stage_award_full_points_to_each_driver(self):
        self.create_championship(drivers=["Ana", "Luis", "Marta"])
        self.create_competition(stages=1)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.finish_competition(
            "Rally Norte",
            ["Ana", "Luis", "Marta"],
            {
                "Ana": ["1:00.000"],
                "Luis": ["1:00.000"],
                "Marta": ["1:10.000"],
            },
        )

        standings = self.service.get_championship_info("Regional")["standings"]
        self.assertEqual(
            [(row["driver"], row["rank"], row["points"]) for row in standings],
            [("Ana", 1, 30), ("Luis", 1, 30), ("Marta", 3, 15)],
        )

    def test_disqualified_driver_loses_bonus_to_next_valid_driver(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"], stages=2)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.finish_competition(
            "Rally Norte",
            ["Ana", "Luis"],
            {
                "Ana": ["0:50.000", "0:50.000"],
                "Luis": ["1:00.000", "1:00.000"],
            },
        )
        self.assertTrue(
            self.service.set_result_status(
                "Rally Norte", "Ana", 2, "Descalificado"
            )[0]
        )

        standings = {row["driver"]: row for row in self.service.get_championship_info("Regional")["standings"]}
        self.assertEqual(standings["Ana"]["points"], 0)
        self.assertEqual(standings["Luis"]["points"], 30)
        self.assertEqual(standings["Luis"]["stage_wins"], 2)

    def test_points_are_not_awarded_until_event_is_finished(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"], stages=1)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.assertTrue(
            self.service.add_time_str("Rally Norte", "Ana", 1, "1:00.000")[0]
        )

        championship = self.service.get_championship_info("Regional")
        self.assertEqual(championship["status"], "En curso")
        self.assertTrue(all(row["points"] == 0 for row in championship["standings"]))

    def test_retired_driver_keeps_stage_win_bonus(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"], stages=2)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.assertTrue(
            self.service.add_time_str("Rally Norte", "Ana", 1, "0:50.000")[0]
        )
        self.assertTrue(
            self.service.add_time_str("Rally Norte", "Luis", 1, "1:00.000")[0]
        )
        self.assertTrue(
            self.service.add_time_str("Rally Norte", "Luis", 2, "1:00.000")[0]
        )
        self.assertTrue(
            self.service.retire_from_rally("Rally Norte", "Ana", 1, True)[0]
        )

        standings = {
            row["driver"]: row
            for row in self.service.get_championship_info("Regional")["standings"]
        }
        self.assertEqual(standings["Ana"]["points"], 23)
        self.assertEqual(standings["Ana"]["stage_wins"], 1)

    def test_configuration_recalculates_points_and_can_close_championship(self):
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"], stages=1)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.finish_competition(
            "Rally Norte",
            ["Ana", "Luis"],
            {"Ana": ["1:00.000"], "Luis": ["1:01.000"]},
        )

        ok, message = self.service.configure_championship(
            "Regional", [10, 5], 0, True
        )
        self.assertTrue(ok, message)
        championship = self.service.get_championship_info("Regional")
        self.assertTrue(championship["manually_finalized"])
        self.assertEqual(championship["points_table"], [10, 5])
        self.assertEqual(championship["standings"][0]["points"], 10)


class ChampionshipParticipationTests(ChampionshipTestCase):
    def test_withdrawal_and_rejoin_preserve_previous_points(self):
        self.create_championship(drivers=["Ana", "Luis"])
        for name in ("Primera", "Segunda"):
            self.create_competition(name, participants=["Ana", "Luis"], stages=1)
            self.assertTrue(
                self.service.add_competition_to_championship("Regional", name)[0]
            )
        self.finish_competition(
            "Primera",
            ["Ana", "Luis"],
            {"Ana": ["1:00.000"], "Luis": ["1:01.000"]},
        )
        self.assertTrue(
            self.service.set_championship_driver_active(
                "Regional", "Ana", 2, False
            )[0]
        )
        self.finish_competition(
            "Segunda",
            ["Ana", "Luis"],
            {"Ana": ["0:59.000"], "Luis": ["1:01.000"]},
        )
        standings = {row["driver"]: row for row in self.service.get_championship_info("Regional")["standings"]}
        self.assertEqual(standings["Ana"]["points"], 30)
        self.assertEqual(standings["Luis"]["points"], 48)

        self.assertTrue(
            self.service.set_championship_driver_active(
                "Regional", "Ana", 2, True
            )[0]
        )
        standings = {row["driver"]: row for row in self.service.get_championship_info("Regional")["standings"]}
        self.assertEqual(standings["Ana"]["points"], 60)
        self.assertEqual(standings["Luis"]["points"], 36)


class ChampionshipExportTests(ChampionshipTestCase):
    def setUp(self):
        super().setUp()
        self.create_championship(drivers=["Ana", "Luis"])
        self.create_competition(participants=["Ana", "Luis"], stages=1)
        self.assertTrue(
            self.service.add_competition_to_championship(
                "Regional", "Rally Norte"
            )[0]
        )
        self.finish_competition(
            "Rally Norte",
            ["Ana", "Luis"],
            {"Ana": ["1:00.000"], "Luis": ["1:01.000"]},
        )

    def test_exports_csv_excel_and_pdf(self):
        csv_path = Path("campeonato.csv")
        excel_path = Path("campeonato.xlsx")
        pdf_path = Path("campeonato.pdf")
        for path in (csv_path, excel_path):
            ok, message = self.service.export_championship("Regional", path)
            self.assertTrue(ok, message)
        ok, message = self.service.export_championship_pdf("Regional", pdf_path)
        self.assertTrue(ok, message)

        csv_text = csv_path.read_text(encoding="utf-8-sig")
        self.assertIn("Rally Norte", csv_text)
        self.assertIn("Ana", csv_text)
        workbook = load_workbook(excel_path, read_only=True)
        self.assertEqual(
            workbook.sheetnames, ["Clasificacion", "Calendario", "Puntuacion"]
        )
        workbook.close()
        self.assertEqual(pdf_path.read_bytes()[:4], b"%PDF")


if __name__ == "__main__":
    unittest.main()
